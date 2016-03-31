import copy
import numbers
import logging
import os

from openpyxl import Workbook

LOG = logging.getLogger("converter-log")
row_count = 1
ws = None


def convert_servers_config(servers_config, nodes):
    """
    Converts the config of servers in the pool
    :param servers_config: F5 servers config for particular pool
    :param nodes: F5 node config to resolve IP of the server
    :return: List of Avi server configs
    """
    server_list = []
    skipped_list = []
    supported_attributes = ['address', 'state']
    for server_name in servers_config.keys():
        server = servers_config[server_name]
        parts = server_name.split(':')
        ip_addr = nodes[parts[0]]["address"]
        port = parts[1] if len(parts) == 2 else 80
        enabled = True
        state = server.get("state", 'enabled')
        if state == "user-down":
            enabled = False
        server_list.append({
            'ip': {
                'addr': ip_addr,
                'type': 'V4'
            },
            'port': port,
            'enabled': enabled
        })
        skipped = [key for key in server.keys() if key not in
                  supported_attributes]
        if skipped:
            skipped_list.append({server_name: skipped})
    return server_list, skipped_list


def get_avi_pool_down_action(action):
    """
    Maps Pool down action from F5 config to Avi Config
    :param action: F5 action string
    :return: Avi action String
    """
    if action == "reset":
        return "POOL_DOWN_ACTION_CLOSE_CONN"
    if action == "reselect":
        return "POOL_DOWN_ACTION_CLOSE_CONN"
    else:
        return "POOL_DOWN_ACTION_CLOSE_CONN"


def convert_pool_config(pool_config, nodes, tenant, monitor_config_list):
    """
    Convert list of pools from F5 config to Avi config
    :param pool_config: F5 pool config
    :param nodes: F5 node list to resolve server IPs
    :param tenant: name of tenant for default objects
    :param monitor_config_list: Avi monitor config list
    :return: List of pools converted to Avi configuration
    """
    pool_list = []
    supported_attr = ['members', 'monitor', 'service-down-action']
    for pool_name in pool_config.keys():
        skipped = []
        f5_pool = pool_config[pool_name]
        if not f5_pool:
            LOG.debug("Empty pool skipped for conversion :"+pool_name)
            add_status_row('pool', None, pool_name, 'skipped', None, None)
            continue
        servers, member_skipped_config = convert_servers_config(
            f5_pool.get("members", {}), nodes)
        sd_action = f5_pool.get("service-down-action", "")
        pd_action = get_avi_pool_down_action(sd_action)
        pool_obj = {
                "name": pool_name,
                "servers": servers,
                "pd_action_type": pd_action
            }
        monitor_names = f5_pool.get("monitor", None)
        default_monitors = {"http": "System-HTTP", "https": "System-HTTPS",
                            "dns": "System-DNS", "tcp": "System-TCP",
                            "udp": "System-UDP", "gateway_icmp": "System-Ping",
                            "icmp": "System-Ping"}
        is_ssl = False
        skipped_monitors = []
        if monitor_names:
            monitors = monitor_names.split(" and ")
            monitor_refs = []
            for monitor in monitors:
                monitor = monitor.strip()
                monitor_obj = [obj for obj in monitor_config_list
                               if obj["name"] == monitor]
                if monitor_obj:
                    monitor_refs.append(monitor_obj[0]["name"])
                    if monitor_obj[0]["type"] == "HEALTH_MONITOR_HTTPS":
                        is_ssl = True
                elif monitor in default_monitors.keys():
                    if monitor == "https":
                        is_ssl = True
                    monitor = tenant+":"+default_monitors[monitor]
                    monitor_refs.append(monitor)
                else:
                    LOG.warning("Skiping non supported monitor: %s for pool %s"
                                % (monitor, pool_name))
                    skipped_monitors.append(monitor)
            pool_obj["health_monitor_refs"] = monitor_refs
        if is_ssl:
            pool_obj["ssl_profile_ref"] = "admin:System-Standard"
        pool_list.append(pool_obj)
        skipped_attr = [key for key in f5_pool.keys() if
                        key not in supported_attr]
        if skipped_attr:
            skipped.append(skipped_attr)
        if member_skipped_config:
            skipped.append(member_skipped_config)
        if skipped_monitors:
            skipped.append({"monitors": skipped_monitors})
        if skipped:
            add_status_row('pool', None, pool_name, 'partial', skipped, pool_obj)
        else:
            add_status_row('pool', None, pool_name, 'successful', skipped, pool_obj)
    return pool_list


def get_profiles_for_vs(profiles, profile_config, tenant):
    """
    Searches for profile refs in converted profile config if not found creates
    default profiles
    :param profiles: profiles in f5 config assigned to VS
    :param profile_config: avi profile config
    :param tenant: tenant name for default profiles
    :return: returns list of profile refs assigned to VS in avi config
    """
    vs_ssl_profile_names = []
    pool_ssl_profile_names = []
    app_profile_names = []
    network_profile_names = []
    if not profiles:
        return []
    for name in profiles.keys():
        ssl_profiles = [obj for obj in profile_config["ssl_profile_list"]
                        if obj['name'] == name]
        if ssl_profiles or name in ["clientssl", "serverssl"]:
            if not ssl_profiles and name == "clientssl":
                vs_ssl_profile_names.append({
                    "profile": tenant + ":System-Standard", "cert": None})
                continue
            if not ssl_profiles and name == "serverssl":
                pool_ssl_profile_names.append({
                    "profile": tenant+":System-Standard", "cert": None})
                continue
            key_cert = [obj for obj in profile_config["ssl_key_cert_list"]
                        if obj['name'] == name]
            key_cert = name if key_cert else None
            profile = profiles.get(name, None)
            context = profile.get("context", None)
            if context == "clientside":
                vs_ssl_profile_names.append({"profile": name, "cert": key_cert})
            elif context == "serverside":
                pool_ssl_profile_names.append(
                    {"profile": name, "cert": key_cert})
        app_profiles = [obj for obj in profile_config["app_profile_list"]
                        if obj['name'] == name]
        if app_profiles or name in ["http", "dns",
                                    "web-acceleration", "http-compression"]:
            if not app_profiles and name in ("http", "fasthttp"):
                app_profile_names.append(tenant+":System-HTTP")
            elif not app_profiles and name == "dns":
                app_profile_names.append(tenant+":System-DNS")
            elif not app_profiles and name == "web-acceleration":
                app_profile = dict()
                http_profile = dict()
                app_profile['name'] = "web-acceleration"
                app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
                cache_config = dict()
                cache_config['query_cacheable'] = True
                cache_config['enabled'] = True
                http_profile["cache_config"] = cache_config
                app_profile["http_profile"] = http_profile
                profile_config["app_profile_list"].append(app_profile)
                app_profile_names.append("web-acceleration")
            elif not app_profiles and name == "http-compression":
                app_profile = dict()
                http_profile = dict()
                app_profile['name'] = "http-compression"
                app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
                compression_profile = dict()
                compression_profile["type"] = "AUTO_COMPRESSION"
                compression_profile["compression"] = True
                http_profile["compression_profile"] = compression_profile
                app_profile["http_profile"] = http_profile
                profile_config["app_profile_list"].append(app_profile)
                app_profile_names.append("http-compression")
            else:
                app_profile_names.append(name)
        network_profiles = [obj for obj in profile_config[
            "network_profile_list"] if obj['name'] == name]
        if network_profiles or name in ["fasthttp", "tcp", "udp"]:
            if not network_profiles and name == "fasthttp":
                network_profile_names.append(tenant+":System-TCP-Proxy")
            if not network_profiles and name == "fastL4":
                network_profile_names.append(tenant+":System-TCP-Fast-Path")
            elif not network_profiles and name == "tcp":
                network_profile_names.append(tenant+":System-TCP-Proxy")
            elif not network_profiles and name == "udp":
                network_profile_names.append(tenant+":System-UDP-Fast-Path")
            else:
                network_profile_names.append(name)

    if not app_profile_names:
        if network_profile_names:
            app_profile_names.append(tenant+":System-L4-Application")
        else:
            app_profile_names.append(tenant+":System-HTTP")

    return vs_ssl_profile_names, pool_ssl_profile_names, \
           app_profile_names, network_profile_names


def clone_pool(pool_name, vs_name, avi_pool_list):
    """
    If pool is shared with other VS pool is cloned for other VS as Avi dose not
    support shared pools with new pool name as <pool_name>-<vs_name>
    :param pool_name: Name of the pool to be cloned
    :param vs_name: Name of the VS for pool to be cloned
    :param avi_pool_list: new pool to be added to this list
    :return: new pool object
    """
    new_pool = None
    for pool in avi_pool_list:
        if pool["name"] == pool_name:
            new_pool = copy.deepcopy(pool)
            break
    if new_pool:
        new_pool["name"] = pool_name+"-"+vs_name
        avi_pool_list.append(new_pool)
        return new_pool["name"]


def add_ssl_to_pool(avi_pool_list, pool_ref, pool_ssl_profiles, tenant):
    """
    F5 serverside SSL need to be added to pool if VS contains serverside SSL
    profile this method add that profile to pool
    :param avi_pool_list: List of pools to search pool object
    :param pool_ref: name of the pool
    :param pool_ssl_profiles: ssl profiles to be added to pool
    :param tenant: tenant name for default object
    """
    if pool_ssl_profiles["profile"] == "serverssl":
        pool_ssl_profiles = tenant+":"+"System-Default-Cert"
    for pool in avi_pool_list:
        if pool_ref == pool["name"]:
            if pool_ssl_profiles["cert"]:
                pool["ssl_key_and_certificate_ref"] = pool_ssl_profiles
            if not pool.get("ssl_profile_ref", None):
                pool["ssl_profile_ref"] = "admin:System-Standard"


def update_service(port, vs, enable_ssl):
    """
    iterates over services of existing vs in converted list to update
    services for port overlapping scenario
    :param port: port for currant VS
    :param vs: VS from converted config list
    :param enable_ssl: value to put in service object
    :return: boolean if service is updated or not
    """
    service_updated = False
    for service in vs['services']:
        port_end = service.get('port_range_end', None)
        if port_end and (service['port'] <= int(port) <= port_end):
            if port not in [1, 65535]:
                new_end = service['port_range_end']
                service['port_range_end'] = int(port)-1
                new_service = {'port': int(port)+1,
                               'port_range_end':new_end,
                               'enable_ssl': enable_ssl}
                vs['services'].append(new_service)
            elif port == 1:
                service['port'] = 2
            elif port == 65535:
                service['port_range_end'] = 65534
            service_updated = True
            break
    return service_updated


def get_service_obj(destination, vs_list, enable_ssl):
    """
    Checks port overlapping scenario for port value 0 in F5 config
    :param destination: IP and Port destination of VS
    :param vs_list: List of existing vs converted to avi config
    :param enable_ssl: value to put in service objects
    :return: List of services for VS
    """
    services_obj = []
    parts = destination.split(':')
    ip_addr = parts[0]
    port = parts[1] if len(parts) == 2 else 80
    vs_dup_ips = [vs for vs in vs_list if vs['ip_address']['addr'] == ip_addr]
    if int(port) > 0:
        for vs in vs_dup_ips:
            service_updated = update_service(port, vs, enable_ssl)
            if service_updated:
                break
        services_obj = [{'port': port, 'enable_ssl': enable_ssl}]
    else:
        used_ports = []
        for vs in vs_dup_ips:
            for service in vs['services']:
                used_ports.append(service['port'])
        if used_ports:
            services_obj = []
            if 65535 not in used_ports:
                used_ports.append(65536)
            used_ports = sorted(used_ports, key=int)
            start = 1
            for i in range(len(used_ports)):
                if start == used_ports[i]:
                    start += 1
                    continue
                end = int(used_ports[i])-1
                services_obj.append({'port': start,
                                     'port_range_end': end,
                                     'enable_ssl': enable_ssl})
                start = int(used_ports[i])+1
        else:
            services_obj = [{'port': 1, 'port_range_end': 65535,
                             'enable_ssl': enable_ssl}]
    return services_obj, ip_addr


def convert_vs_config(vs_config, vs_state, tenant,
                      avi_pool_list, profile_config):
    """
    F5 virtual server object conversion to Avi VS object
    :param vs_config: F5 virtual server config list
    :param vs_state: state of new VS to be created in Avi
    :param tenant: tenant name for default object
    :param avi_pool_list: List of pools to handle shared pool scenario
    :param profile_config: Avi profile config for profiles referenced in vs

    :return: List of Avi VS configs
    """
    vs_list = []
    supported_attr = ['profiles', 'destination', 'pool']
    for vs_name in vs_config.keys():
        f5_vs = vs_config[vs_name]
        skipped = [key for key in f5_vs.keys() if
                        key not in supported_attr]
        enabled = (vs_state == 'enable')
        ssl_vs, ssl_pool, app_prof, ntwk_prof = get_profiles_for_vs(f5_vs.get(
            "profiles", None), profile_config, tenant)
        enable_ssl = False
        if ssl_vs:
            enable_ssl = True
            if app_prof[0] == (tenant+':System-HTTP'):
                app_prof[0] = tenant+':System-Secure-HTTP'
        destination = f5_vs["destination"]
        services_obj, ip_addr = get_service_obj(destination, vs_list,
                                                enable_ssl)
        pool_ref = f5_vs.get("pool", None)
        if pool_ref:
            shared_vs = [obj for obj in vs_list
                         if obj.get("pool_ref", "") == pool_ref]
            if shared_vs:
                pool_ref = clone_pool(pool_ref, vs_name, avi_pool_list)
            if ssl_pool:
                add_ssl_to_pool(avi_pool_list, pool_ref, ssl_pool[0], tenant)

        vs_obj = {
            'name': vs_name,
            'type': 'VS_TYPE_NORMAL',
            'ip_address': {
                'addr': ip_addr,
                'type': 'V4'
            },
            'enabled': enabled,
            'services': services_obj,
            'application_profile_ref': app_prof[0],
            'pool_ref': pool_ref
        }
        if ntwk_prof:
            vs_obj['network_profile_ref'] = ntwk_prof[0]
        if enable_ssl:
            default_cert = tenant+":"+"System-Default-Cert"
            vs_obj['ssl_profile_name'] = ssl_vs[0]["profile"]
            if ssl_vs[0]["cert"]:
                vs_obj['ssl_key_and_certificate_refs'] = [ssl_vs[0]["cert"]]
            else:
                vs_obj['ssl_key_and_certificate_refs'] = [default_cert]

        vs_list.append(vs_obj)
        if skipped:
            add_status_row('virtual', None, vs_name, 'partial', skipped, vs_obj)
        else:
            add_status_row('virtual', None, vs_name, 'successful',
                           skipped, vs_obj)

    return vs_list


def get_defaults(monitor_type, f5_monitor, monitor_config):
    """
    Monitor can have inheritance used by attribute defaults-from in F5
    configuration this method recursively gets all the attributes from the
    default objects and forms complete object
    :param monitor_type: Monitor type
    :param f5_monitor: F5 monitor object
    :param monitor_config: List of F5 monitor configs
    :return:
    """
    parent_name = f5_monitor.get("defaults-from", None)
    if parent_name:
        parent_monitor = monitor_config.get(monitor_type+" "+parent_name, None)
        if parent_monitor:
            parent_monitor = get_defaults(monitor_type,
                                          parent_monitor, monitor_config)
            parent_monitor = copy.deepcopy(parent_monitor)
            parent_monitor.update(f5_monitor)
            f5_monitor = parent_monitor
    return f5_monitor


def convert_monitor_entity(monitor_type, name, f5_monitor):
    """
    Conversion of single F5 monitor object to Avi health monitor object
    :param monitor_type: Health monitor type
    :param name: name of health monitor
    :param f5_monitor: F5 monitor config object
    :return: Avi monitor config object
    """
    supported_attributes = ["timeout", "interval", "time-until-up",
                            "description"]
    skipped = [key for key in f5_monitor.keys() if
              key not in supported_attributes]
    timeout = int(f5_monitor.get("timeout", 16))
    interval = int(f5_monitor.get("interval", 5))
    time_until_up = int(f5_monitor.get("time-until-up", 1))
    successful_checks = int(timeout/interval)
    if time_until_up > 0:
        failed_checks = int(time_until_up/interval)
        failed_checks = 1 if failed_checks == 0 else failed_checks
    else:
        failed_checks = 1
    description = f5_monitor.get("description", None)
    monitor_dict = dict()
    monitor_dict["name"] = name
    monitor_dict["receive_timeout"] = interval-1
    monitor_dict["failed_checks"] = failed_checks
    monitor_dict["send_interval"] = interval
    monitor_dict["successful_checks"] = successful_checks
    if description:
        monitor_dict["description"] = description

    if monitor_type == "http":
        monitor_dict["type"] = "HEALTH_MONITOR_HTTP"
        monitor_dict["http_monitor"] = {
            "http_request": "HEAD / HTTP/1.0",
            "http_response_code": [
                {"code": "HTTP_2XX"}, {"code": "HTTP_3XX"}
            ]}
    elif monitor_type == "https":
        monitor_dict["type"] = "HEALTH_MONITOR_HTTPS"
        monitor_dict["https_monitor"] = {
            "http_request": "HEAD / HTTP/1.0",
            "http_response_code": [
                {"code": "HTTP_2XX"}, {"code": "HTTP_3XX"}
            ]}
    elif monitor_type == "dns":
        skipped = [key for key in skipped if key not in ["accept-rcode"]]
        accept_rcode = f5_monitor.get("accept-rcode", None)
        if accept_rcode and accept_rcode == "no-error":
            rcode = "RCODE_NO_ERROR"
        else:
            rcode = "RCODE_ANYTHING"
        monitor_dict["type"] = "HEALTH_MONITOR_DNS"
        dns_monitor = dict()
        dns_monitor["rcode"] = rcode
        dns_monitor["query_name"] = f5_monitor.get("qname", None)
        dns_monitor["qtype"] = "DNS_QUERY_TYPE"
        monitor_dict["dns_monitor"] = dns_monitor
    elif monitor_type == "tcp":
        tcp_attr = ["destination", "send", "recv"]
        skipped = [key for key in skipped if key not in tcp_attr]
        destination = f5_monitor.get("destination", "*:*")
        dest_str = destination.split(":")
        if len(dest_str) > 1 and isinstance(dest_str[1], numbers.Integral):
            monitor_dict["monitor_port"] = dest_str[1]
        monitor_dict["type"] = "HEALTH_MONITOR_TCP"
        request = f5_monitor.get("send", None)
        response = f5_monitor.get("recv", None)
        if request or response:
            tcp_monitor = {"tcp_request": request, "tcp_response": response}
            monitor_dict["tcp_monitor"] = tcp_monitor
    elif monitor_type == "udp":
        udp_attr = ["destination", "send", "recv"]
        skipped = [key for key in skipped if key not in udp_attr]
        destination = f5_monitor.get("destination", "*:*")
        dest_str = destination.split(":")
        if len(dest_str) > 1 and isinstance(dest_str[1], numbers.Integral):
            monitor_dict["monitor_port"] = dest_str[1]
        monitor_dict["type"] = "HEALTH_MONITOR_UDP"
        request = f5_monitor.get("send", None)
        response = f5_monitor.get("recv", None)
        if request or response:
            tcp_monitor = {"udp_request": request, "udp_response": response}
            monitor_dict["tcp_monitor"] = tcp_monitor
    elif monitor_type in ["gateway-icmp", "icmp"]:
        monitor_dict["type"] = "HEALTH_MONITOR_PING"
    elif monitor_type == "external":
        ext_attr = ["run", "args", "user-defined"]
        skipped = [key for key in skipped if key not in ext_attr]
        monitor_dict["type"] = "HEALTH_MONITOR_EXTERNAL"
        ext_monitor = {
            "command_code": f5_monitor.get("run", None),
            "command_parameters": f5_monitor.get("args", None),
            "command_variables": f5_monitor.get("user-defined", None)
        }
        monitor_dict["external_monitor"] = ext_monitor
    return monitor_dict, skipped


def convert_monitor_config(monitor_config):
    """
    Convert F5 monitor config dict to Avi health monitor config list
    :param monitor_config: F5 monitor config dict
    :return: List of Avi health monitor objects
    """
    monitor_list = []

    supported_types = ["http", "https", "dns", "external", "tcp", "udp",
                       "gateway-icmp", "icmp"]
    for key in monitor_config.keys():
        monitor_type, name = key.split(" ")
        if monitor_type not in supported_types:
            LOG.debug("Monitor type not supported by Avi : "+name)
            add_status_row('monitor', monitor_type, name, 'skipped', None, None)
            continue
        f5_monitor = monitor_config[key]
        if not f5_monitor:
            add_status_row('monitor', monitor_type, name, 'skipped', None, None)
            continue
        f5_monitor = get_defaults(monitor_type, f5_monitor, monitor_config)
        avi_monitor, skipped = convert_monitor_entity(
            monitor_type, name, f5_monitor)
        if skipped:
            add_status_row('monitor', monitor_type, name, 'partial',
                           skipped, avi_monitor)
        else:
            add_status_row('monitor', monitor_type, name, 'successful',
                           None, avi_monitor)
        monitor_list.append(avi_monitor)
    return monitor_list


def get_key_cert_obj(name, key_file_name, cert_file_name, folder_path, option):
    """
    Read key and cert files from given location and construct avi
    SSLKeyAndCertificate objects
    :param name: SSLKeyAndCertificate object name
    :param key_file_name: key file name
    :param cert_file_name: cert file name
    :param folder_path: location of key and cert files
    :param option: api-upload or cli-file both requires different
    object structure
    :return:SSLKeyAndCertificate object
    """
    key = None
    cert = None
    folder_path = folder_path+os.path.sep
    try:
        key_file = open(folder_path+key_file_name, "r")
        key = key_file.read()
    except:
        LOG.error("Error to read file" + folder_path+key_file_name)
    try:
        cert_file = open(folder_path+cert_file_name, "r")
        cert = cert_file.read()
    except:
        LOG.error("Error to read file" + folder_path+cert_file_name)
    ssl_kc_obj = None
    if key and cert:
        if option == "cli-upload":
            cert = {"certificate": cert}
        ssl_kc_obj = {
                'name': name,
                'key': key,
                'certificate': cert,
                'key_passphrase': ''
            }
    return ssl_kc_obj


def update_with_default_profile(profile_type, profile, profile_config):
    """
    Profiles can have inheritance used by attribute defaults-from in F5
    configuration this method recursively gets all the attributes from the
    default objects and forms complete object
    :param profile_type: type of profile
    :param profile: currant profile object
    :param profile_config: F5 profile config dict
    :return: Complete profile with updated attributes from defaults
    """
    parent_name = profile.get("defaults-from", None)
    if parent_name:
        parent_profile = profile_config.get(profile_type + " " +
                                            parent_name, None)
        if parent_profile:
            parent_profile = get_defaults(profile_type,
                                          parent_profile, profile_config)
            parent_profile = copy.deepcopy(parent_profile)
            parent_profile.update(profile)
            profile = parent_profile
    return profile


def convert_profile_config(profile_config, certs_location, option):
    """
    Converts F5 profiles to equivalent Avi profiles
    :param profile_config: F5 Profile config dict
    :param certs_location: location of cert and key file location
    :param option: api-upload or cli-file both requires different
    object structure
    :return:
    """
    ssl_key_cert_list = []
    app_profile_list = []
    ssl_profile_list = []
    pki_profile_list = []
    string_group = []
    network_profile_list = []
    for key in profile_config.keys():
        converted_objs = []
        profile_type, name = key.split(" ")
        profile = profile_config[key]
        skipped = profile.keys()
        profile = update_with_default_profile(profile_type,
                                              profile, profile_config)
        if profile_type in ("client-ssl", "server-ssl"):
            supported_attr = ["cert-key-chain", "cert", "key",
                              "ciphers", "unclean-shutdown",
                              "crl-file", "ca-file"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            cert_obj = profile.get("cert-key-chain", None)
            key_file = None
            cert_file = None
            key_cert_obj = None
            if cert_obj:
                key_file = cert_obj["default"]["key"]
                cert_file = cert_obj["default"]["cert"]
            elif profile.get("cert", None):
                cert_file = profile["cert"]
                key_file = profile["key"]
            if key_file and cert_file:
                key_cert_obj = get_key_cert_obj(
                    name, key_file, cert_file, certs_location, option)
            if key_cert_obj:
                ssl_key_cert_list.append(key_cert_obj)
                converted_objs.append({'key_cert': key_cert_obj})
            ciphers = profile.get('ciphers', 'DEFAULT')
            ciphers = 'AES:3DES:RC4' if ciphers == 'DEFAULT' else ciphers
            ssl_profile = dict()
            ssl_profile['name'] = name
            ssl_profile['accepted_ciphers'] = ciphers
            close_notify = profile.get('unclean-shutdown', None)
            if close_notify and close_notify == 'enabled':
                ssl_profile['send_close_notify'] = True
            else:
                ssl_profile['send_close_notify'] = False
            ssl_profile_list.append(ssl_profile)
            converted_objs.append({'ssl_profile': ssl_profile})
            options = profile.get("options", "")
            options = options.keys()+options.values()
            accepted_versions = []
            if "no-tlsv1" not in options:
                accepted_versions.append({"type": "SSL_VERSION_TLS1"})
            if "no-tlsv1.1" not in options:
                accepted_versions.append({"type": "SSL_VERSION_TLS1_1"})
            if "no-tlsv1.2" not in options:
                accepted_versions.append({"type": "SSL_VERSION_TLS1_2"})
            if accepted_versions:
                ssl_profile["accepted_versions"] = accepted_versions
            if options:
                skipped_options = [key for key in options if key not in [
                    "no-tlsv1", "no-tlsv1.1", "no-tlsv1.2", None]]
                skipped.append({"Unsupported options": skipped_options})

            crl_file_name = profile.get('crl-file', None)
            ca_file_name = profile.get('ca-file', None)
            crl_file_name = None if crl_file_name == 'none' else crl_file_name
            ca_file_name = None if ca_file_name == 'none' else ca_file_name
            if not ca_file_name and crl_file_name:
                LOG.warning("Skipped PKI profile for profile %s "
                            "because of missing CA file" % name)

            elif ca_file_name:
                pki_profile = dict()
                file_path = certs_location+os.path.sep+ca_file_name
                pki_profile["name"] = name
                error = False
                try:
                    ca_file = open(file_path, "r")
                    ca = ca_file.read()
                    pki_profile["ca_certs"] = [{'certificate': ca}]
                except:
                    LOG.error("Error to read file" +
                              certs_location+os.path.sep+ca_file_name)
                    error = True
                if crl_file_name:
                    file_path = certs_location+os.path.sep+crl_file_name
                    try:
                        crl_file = open(file_path, "r")
                        crl = crl_file.read()
                        pki_profile["crls"] = [{'body': crl}]
                    except:
                        LOG.error("Error to read file" +
                                  certs_location+os.path.sep+crl_file_name)
                        error = True
                if not error:
                    pki_profile_list.append(pki_profile)
                    converted_objs.append({'pki_profile': pki_profile})
        elif profile_type == 'http':
            supported_attr = ["description", "insert-xforwarded-for",
                              "enforcement", "xff-alternative-names"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            app_profile = dict()
            app_profile['name'] = name
            app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
            app_profile['description'] = profile.get('description', None)
            http_profile = dict()
            http_profile['x_forwarded_proto_enabled'] = \
                profile.get('insert-xforwarded-for', False)
            http_profile['xff_alternate_name'] = \
                profile.get('xff-alternative-names', None)
            enforcement = profile.get('enforcement', None)
            if enforcement:
                header_size = enforcement.get('max-header-size', 49152)
                http_profile['client_max_header_size'] = int(header_size)/1024
                enf_skipped = [key for key in enforcement.keys()
                               if key not in ["max-header-size"]]
                skipped.append({"enforcement": enf_skipped})
            app_profile["http_profile"] = http_profile
            app_profile_list.append(app_profile)
            converted_objs.append({'app_profile': app_profile})
        elif profile_type == 'dns':
            app_profile = dict()
            app_profile['name'] = name
            app_profile['type'] = 'APPLICATION_PROFILE_TYPE_DNS'
            app_profile['description'] = profile.get('description', None)
            app_profile_list.append(app_profile)
            converted_objs.append({'app_profile': app_profile})
        elif profile_type == 'web-acceleration':
            supported_attr = ["description", "cache-object-min-size",
                              "cache-max-age", "cache-object-max-size",
                              "cache-insert-age-header"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            app_profile = dict()
            app_profile['name'] = name
            app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
            app_profile['description'] = profile.get('description', None)
            cache_config = dict()
            cache_config['min_object_size'] = profile.get(
                'cache-object-min-size', 100)
            cache_config['query_cacheable'] = True
            cache_config['max_object_size'] = profile.get(
                'cache-object-max-size', 4194304)
            age_header = profile.get('cache-insert-age-header', 'disabled')
            if age_header == 'enabled':
                cache_config['age_header'] = True
            else:
                cache_config['age_header'] = False
            cache_config['enabled'] = True
            cache_config['default_expire'] = profile.get('cache-max-age', 600)
            # max_entities = profile.get('cache-max-entries', 0)
            # cache_config['max_cache_size'] = \
            #     (int(max_entities) * int(cache_config['max_object_size']))
            http_profile = dict()
            http_profile["cache_config"] = cache_config
            app_profile["http_profile"] = http_profile
            app_profile_list.append(app_profile)
            converted_objs.append({'app_profile': app_profile})
        elif profile_type == 'http-compression':
            supported_attr = ["description", "content-type-include",
                              "keep-accept-encoding"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            app_profile = dict()
            app_profile['name'] = name
            app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
            app_profile['description'] = profile.get('description', None)
            compression_profile = dict()
            compression_profile["type"] = "AUTO_COMPRESSION"
            compression_profile["compression"] = True
            encoding = profile.get("keep-accept-encoding", "disable")
            if encoding == "disable":
                encoding = True
            else:
                encoding = False
            compression_profile["remove_accept_encoding_header"] = encoding
            content_type = profile.get("content-type-include", "")
            if content_type:
                content_types = content_type.keys()+content_type.values()
                sg_obj = {"name": name+"-content_type"}
                uris = []
                for content_type in content_types:
                    uri = {"str": content_type}
                    uris.append(uri)
                sg_obj["uris"] = uris
                string_group.append(sg_obj)
                converted_objs.append({'string_group': sg_obj})
                compression_profile["compressible_content_ref"] = \
                    name + "-content_type"
            http_profile = dict()
            http_profile["compression_profile"] = compression_profile
            app_profile["http_profile"] = http_profile
            app_profile_list.append(app_profile)
            converted_objs.append({'app_profile': app_profile})
        elif profile_type == 'fastl4':
            supported_attr = ["description", "explicit-flow-migration",
                              "idle-timeout", "software-syn-cookie"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            syn_protection = (profile.get("software-syn-cookie",
                                          None) == 'enabled')
            timeout = profile.get("idle-timeout", 0)
            ntwk_profile = {
                  "profile": {
                        "tcp_fast_path_profile": {
                          "session_idle_timeout": timeout,
                          "enable_syn_protection": syn_protection
                        },
                        "type": "PROTOCOL_TYPE_TCP_FAST_PATH"
                  },
                  "name": name
                }
            app_profile = dict()
            app_profile['name'] = name
            app_profile['type'] = 'APPLICATION_PROFILE_TYPE_L4'
            explicit_tracking = profile.get("explicit-flow-migration",None)
            l4_profile = {"rl_profile": {
                    "client_ip_connections_rate_limit": {
                        "explicit_tracking": (explicit_tracking == 'enabled')
                    }}
            }
            app_profile['dos_rl_profile'] = l4_profile
            app_profile_list.append(app_profile)
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'fasthttp':
            supported_attr = ["description", "receive-window-size",
                              "idle-timeout"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            receive_window = profile.get("receive-window-size", 64)
            if not (32 <= int(receive_window) <= 65536):
                receive_window = 64
            timeout = profile.get("idle-timeout", 0)
            ntwk_profile = {
                "profile": {
                    "tcp_proxy_profile": {
                        "receive_window": receive_window,
                        "idle_connection_timeout": timeout
                    },
                    "type": "PROTOCOL_TYPE_TCP_PROXY"
                },
                "name": name
            }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'tcp':
            supported_attr = ["description", " idle-timeout"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            timeout = profile.get("idle-timeout", 0)
            ntwk_profile = {
                  "profile": {
                        "tcp_proxy_profile": {
                          "session_idle_timeout": timeout
                        },
                        "type": "PROTOCOL_TYPE_TCP_PROXY"
                  },
                  "name": name
                }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'udp':
            supported_attr = ["description", "idle-timeout",
                              "datagram-load-balancing"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            per_pkt = profile.get("datagram-load-balancing", 'disabled')
            timeout = profile.get("idle-timeout", 0)
            ntwk_profile = {
                "profile": {
                    "type": "PROTOCOL_TYPE_UDP_FAST_PATH",
                    "udp_fast_path_profile": {
                        "per_pkt_loadbalance": (per_pkt == 'enabled'),
                        "session_idle_timeout": timeout
                    }
                },
                "name": "System-UDP-Fast-Path"
            }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        else:
            LOG.warning("Not supported profile type: %s" % profile_type)
            add_status_row('profile', profile_type, name, 'skipped', None, None)
            continue
        if skipped:
            add_status_row('profile', profile_type, name, 'partial', skipped, converted_objs)
        else:
            add_status_row('profile', profile_type, name, 'successful', skipped, converted_objs)
    avi_profiles = dict()
    avi_profiles["ssl_key_cert_list"] = ssl_key_cert_list
    avi_profiles["app_profile_list"] = app_profile_list
    avi_profiles["ssl_profile_list"] = ssl_profile_list
    avi_profiles["pki_profile_list"] = pki_profile_list
    avi_profiles["network_profile_list"] = network_profile_list
    return avi_profiles, string_group


def add_status_row(f5_type, f5_sub_type, f5_id, status, skipped_params,
                   avi_object):
    global row_count
    global ws
    ws['A'+str(row_count)] = f5_type
    ws['B'+str(row_count)] = f5_sub_type
    ws['C'+str(row_count)] = f5_id
    ws['D'+str(row_count)] = status
    ws['E'+str(row_count)] = str(skipped_params)
    ws['F'+str(row_count)] = str(avi_object)
    row_count += 1


def convert_to_avi_dict(f5_config_dict, output_file_path,
                        vs_state, certs_location, tenant, option):
    global row_count
    global ws
    wb = Workbook()
    ws = wb.active
    ws['A'+str(row_count)] = 'F5 type'
    ws['B'+str(row_count)] = 'F5 SubType'
    ws['C'+str(row_count)] = 'F5 ID'
    ws['D'+str(row_count)] = 'Status'
    ws['E'+str(row_count)] = 'Skipped settings'
    ws['F'+str(row_count)] = 'Avi Object'
    row_count += 1
    avi_config_dict = {}
    monitor_config_list = convert_monitor_config(f5_config_dict.get(
        "monitor", {}))
    del f5_config_dict["monitor"]
    avi_config_dict["HealthMonitor"] = monitor_config_list
    LOG.debug("Converted health monitors")
    avi_pool_list = convert_pool_config(f5_config_dict.get("pool", {}),
                                        f5_config_dict.get("node", {}), tenant,
                                        monitor_config_list)
    del f5_config_dict["pool"]
    avi_config_dict["Pool"] = avi_pool_list
    LOG.debug("Converted pools")
    f5_profile_dict = f5_config_dict.get("profile", {})
    avi_profiles, string_group = convert_profile_config(
        f5_profile_dict, certs_location, option)
    del f5_config_dict["profile"]
    avi_config_dict["SSLKeyAndCertificate"] = avi_profiles["ssl_key_cert_list"]
    avi_config_dict["SSLProfile"] = avi_profiles["ssl_profile_list"]
    avi_config_dict["PKIProfile"] = avi_profiles["pki_profile_list"]
    avi_config_dict["ApplicationProfile"] = avi_profiles["app_profile_list"]
    avi_config_dict["NetworkProfile"] = avi_profiles["network_profile_list"]
    avi_config_dict["StringGroup"] = string_group
    LOG.debug("Converted ssl profiles")
    avi_vs_list = convert_vs_config(f5_config_dict.get("virtual", {}), vs_state,
                                    tenant, avi_pool_list, avi_profiles)
    avi_config_dict["VirtualService"] = avi_vs_list
    del f5_config_dict["virtual"]
    LOG.debug("Converted VS")
    for f5_type in f5_config_dict.keys():
        f5_obj = f5_config_dict[f5_type]
        for key in f5_obj.keys():
            sub_type = None
            if ' ' in key:
                sub_type, key = key.rsplit(' ', 1)
            add_status_row(f5_type, sub_type, key, 'skipped', None, None)
    wb.save(filename = output_file_path+os.path.sep+"ConversionStatus.xls")
    return avi_config_dict
